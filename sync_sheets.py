import re
import os
from datetime import datetime

import gspread
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

CREDENTIALS_PATH = os.path.join(os.path.dirname(__file__), "credentials", "credentials.json")
SPREADSHEET_ID = "1kuHPwbQ1LGlNPqBHNuKrVNhUpaOi62kB8uqw8KJu1EY"
SHEET_NAME = "М/Сервис"

COMPETITOR_COLUMNS = {
    "Cepco": 29,
    "Hard Workers": 30,
    "Display мастер": 31,
    "Kiber Centre": 32,
    "Fixed one": 33,
    "Apple Pro": 34,
    "Brobrolab": 35,
    "Dabro": 36,
    "Modmac": 37,
    "Planet iPhone": 38,
    "Mos display": 39,
    "iSupport": 40,
    "jabuka": 41,
    "Apple Pie": 42,
    "Станислава": 43,
}

IPHONE_MODEL_TO_SHEET = {
    "iPhone X": "X",
    "iPhone XR": "XR",
    "iPhone XS": "XS",
    "iPhone XS Max": "XS Max",
    "iPhone 11": "11",
    "iPhone 11 Pro": "11 Pro",
    "iPhone 11 Pro Max": "11 Pro Max",
    "iPhone 12 mini": "12 Mini",
    "iPhone 12": "12",
    "iPhone 12 Pro": "12 Pro",
    "iPhone 12 Pro Max": "12 Pro Max",
    "iPhone 13 mini": "13 Mini",
    "iPhone 13": "13",
    "iPhone 13 Pro": "13 Pro",
    "iPhone 13 Pro Max": "13 Pro Max",
    "iPhone 14": "14",
    "iPhone 14 Plus": "14 Plus",
    "iPhone 14 Pro": "14 Pro",
    "iPhone 14 Pro Max": "14 Pro Max",
    "iPhone 15": "15",
    "iPhone 15 Plus": "15 Plus",
    "iPhone 15 Pro": "15 Pro",
    "iPhone 15 Pro Max": "15 Pro Max",
    "iPhone 16": "16",
    "iPhone 16E": "16E",
    "iPhone 16 Plus": "16 Plus",
    "iPhone 16 Pro": "16 Pro",
    "iPhone 16 Pro Max": "16 Pro Max",
    "iPhone 17": "17",
    "iPhone 17 Air": "17 Air",
    "iPhone 17 Pro": "17 Pro",
    "iPhone 17 Pro Max": "17 Pro Max",
}

IPHONE_REPAIR_TO_SHEET = {
    "Замена дисплея": "Замена дисплея",
    "Замена заднего стекла": "Замена заднего стекла",
    "Замена аккумулятора": "Замена аккумулятора",
    "Замена основной камеры": "Замена камеры| Задняя камера",
    "Замена разъема зарядки": "Замена разъема зарядки",
    "Замена стекла": "Замена стекла",
    "Замена разговорного микрофона": "Замена разговорного микрофона",
    "Замена слухового динамика": "Замена слухового динамика",
    "Замена полифонического динамика": "Замена полифонического динамика",
    "Замена динамика": None,
    "Неизвестно": None,
}

MACBOOK_MODEL_TO_SHEET = {}
MACBOOK_REPAIR_TO_SHEET = {
    "Замена матрицы": "Замена матрицы",
    "Замена дисплея (матрицы)": "Замена матрицы",
    "Замена дисплея в сборе": "Замена дисплея в сборе",
    "Замена экрана в сборе": "Замена дисплея в сборе",
    "Замена аккумулятора": "Замена аккумулятора",
    "Замена клавиатуры": "Замена клавиатуры",
    "Замена SSD": None,
    "Замена трекпада": "Замена тачпада ",
    "Замена динамика": None,
    "Замена вентилятора": "Чистка и замена термопасты Macbook",
    "Замена топкейса": None,
    "Ремонт после попадания воды": "Чистка после попадания жидкости Macbook",
}


def _build_macbook_model_map():
    models = [
        ('MacBook Air 13" M4', "Air 13"),
        ('MacBook Air 13" M3', "Air 13"),
        ('MacBook Air 13" M2', "Air 13"),
        ('MacBook Air 13" M1', "Air 13"),
        ('MacBook Air 13" 2018-2020', "Air 13"),
        ('MacBook Air 11-13" 2010-2017', "Air 11"),
        ('MacBook Air 15" M4', "Air 15"),
        ('MacBook Air 15" M3', "Air 15"),
        ('MacBook Air 15" M2', "Air 15"),
        ('MacBook Pro 14" M4', "Pro 14"),
        ('MacBook Pro 14" M3', "Pro 14"),
        ('MacBook Pro 14" M2', "Pro 14"),
        ('MacBook Pro 14" M1', "Pro 14"),
        ('MacBook Pro 16" M4', "Pro 16"),
        ('MacBook Pro 16" M3', "Pro 16"),
        ('MacBook Pro 16" M2', "Pro 16"),
        ('MacBook Pro 16" M1', "Pro 16"),
        ('MacBook Pro 16" Intel', "Pro 16"),
        ('MacBook Pro 15" 2018-2019', "Pro 15"),
        ('MacBook Pro 15" 2016-2017', "Pro 15"),
        ('MacBook Pro 15" 2012-2015', "Pro 15"),
        ('MacBook Pro 13" M2', "Pro 13"),
        ('MacBook Pro 13" M1', "Pro 13"),
        ('MacBook Pro 13" 2018-2020', "Pro 13"),
        ('MacBook Pro 13" 2016-2017', "Pro 13"),
        ('MacBook Pro 13" 2012-2015', "Pro 13"),
        ('MacBook 12"', "12 | 12"),
    ]
    for full_name, sheet_name in models:
        MACBOOK_MODEL_TO_SHEET[full_name] = sheet_name

    models_with_years = [
        ('MacBook Air 13" M1 2020', "Air 13"),
        ('MacBook Air 13" M2 2022', "Air 13"),
        ('MacBook Air 15" M2 2023', "Air 15"),
        ('MacBook Air 13" M3 2024', "Air 13"),
        ('MacBook Air 15" M3 2024', "Air 15"),
        ('MacBook Pro 13" M1 2020', "Pro 13"),
        ('MacBook Pro 13" M2 2022', "Pro 13"),
        ('MacBook Pro 14" M1 Pro 2021', "Pro 14"),
        ('MacBook Pro 16" M1 Pro 2021', "Pro 16"),
        ('MacBook Pro 14" M2 Pro 2023', "Pro 14"),
        ('MacBook Pro 16" M2 Pro 2023', "Pro 16"),
        ('MacBook Pro 14" M3 2023', "Pro 14"),
        ('MacBook Pro 16" M3 Pro 2023', "Pro 16"),
        ('MacBook Pro 14" M4 2024', "Pro 14"),
        ('MacBook Pro 16" M4 2024', "Pro 16"),
    ]
    for full_name, sheet_name in models_with_years:
        if full_name not in MACBOOK_MODEL_TO_SHEET:
            MACBOOK_MODEL_TO_SHEET[full_name] = sheet_name


_build_macbook_model_map()


def _row_key(g, h, qi, j, l=""):
    if l:
        return f"{g}|{h}|{qi}|{j}|{l}"
    return f"{g}|{h}|{qi}|{j}"


def _load_sheet_index(ws):
    all_vals = ws.get_all_values()
    index = {}
    a_number_index = {}
    for rnum, rv in enumerate(all_vals, 1):
        g = rv[6].strip() if len(rv) > 6 else ""
        h = rv[7].strip() if len(rv) > 7 else ""
        qi = rv[8].strip() if len(rv) > 8 else ""
        j = rv[9].strip() if len(rv) > 9 else ""
        l = rv[11].strip() if len(rv) > 11 else ""
        if not g and not h and not j:
            continue
        key = _row_key(g, h, qi, j)
        if key not in index:
            index[key] = rnum
        if g.lower() == "macbook" and l:
            for a_num in re.findall(r"A\d{4}", l, re.IGNORECASE):
                akey = (h, qi, j, a_num.upper())
                if akey not in a_number_index:
                    a_number_index[akey] = rnum
    return index, a_number_index


def _map_quality(quality: str) -> str:
    if quality == "AASP":
        return "AASP"
    if quality == "OEM":
        return "OEM"
    return ""


def _find_sheet_repair(repair: str, repair_map: dict, sheet_repairs: list) -> str | None:
    if repair in repair_map:
        mapped = repair_map[repair]
        if mapped is None:
            return None
        return mapped
    rl = repair.lower().strip()
    for sheet_repair in sheet_repairs:
        sl = sheet_repair.lower().strip()
        if rl in sl or sl in rl:
            return sheet_repair
    return None


def _match_row(device_type, sheet_model, sheet_repair, sheet_quality, index, a_number_index=None, a_number=None):
    if device_type == "Macbook" and a_number_index and a_number:
        a_upper = a_number.upper().strip()
        qualities = [sheet_quality] if sheet_quality in ("AASP", "OEM") else ["OEM", "AASP", "-", "HQ"]
        for q in qualities:
            akey = (sheet_model, q, sheet_repair, a_upper)
            if akey in a_number_index:
                return a_number_index[akey]
    keys = []
    if sheet_quality == "AASP":
        keys.append(_row_key(device_type, sheet_model, "AASP", sheet_repair))
    elif sheet_quality == "OEM":
        keys.append(_row_key(device_type, sheet_model, "OEM", sheet_repair))
    else:
        keys.append(_row_key(device_type, sheet_model, "OEM", sheet_repair))
        keys.append(_row_key(device_type, sheet_model, "AASP", sheet_repair))
        keys.append(_row_key(device_type, sheet_model, "-", sheet_repair))
        keys.append(_row_key(device_type, sheet_model, "HQ", sheet_repair))
    for key in keys:
        if key in index:
            return index[key]
    return None


def _build_updates(data: list[dict], index: dict, sheet_repairs: dict, model_map: dict, repair_map: dict, a_number_index: dict = None) -> list[dict]:
    from normalize import MACBOOK_A_NUMBER_MAP
    model_to_a_numbers = {}
    for a_num, full_model in MACBOOK_A_NUMBER_MAP.items():
        if full_model not in model_to_a_numbers:
            model_to_a_numbers[full_model] = []
        model_to_a_numbers[full_model].append(a_num)

    updates = []
    for d in data:
        model = d.get("model", "")
        repair = d.get("repair", "")
        price = d.get("price", 0)
        competitor = d.get("competitor", "")
        quality_raw = d.get("quality", "")
        device_type = d.get("device_type", "")
        marking = d.get("marking", "")
        if not device_type:
            if "iphone" in model.lower() or "айфон" in model.lower():
                device_type = "iPhone"
            elif "macbook" in model.lower():
                device_type = "Macbook"
            else:
                continue

        sheet_model = model_map.get(model)
        if not sheet_model:
            continue

        dt_key = device_type
        if dt_key == "iphone":
            dt_key = "iPhone"
        elif dt_key == "macbook":
            dt_key = "Macbook"

        sheet_repair = _find_sheet_repair(repair, repair_map, sheet_repairs.get(dt_key, []))
        if not sheet_repair:
            continue

        sheet_quality = _map_quality(quality_raw)
        col = COMPETITOR_COLUMNS.get(competitor)
        if not col:
            continue

        a_number = marking if marking else None
        if not a_number and dt_key == "Macbook":
            a_numbers = model_to_a_numbers.get(model, [])
            a_number = a_numbers[0] if a_numbers else None

        row = _match_row(dt_key, sheet_model, sheet_repair, sheet_quality, index, a_number_index, a_number)
        if not row:
            continue

        updates.append({
            "row": row, "col": col, "price": price,
            "competitor": competitor, "model": model,
            "repair": repair, "quality": quality_raw,
            "sheet_model": sheet_model, "sheet_repair": sheet_repair,
            "sheet_quality": sheet_quality, "device_type": dt_key,
        })

    return updates


def _deduplicate_updates(updates: list[dict]) -> list[dict]:
    best = {}
    for u in updates:
        key = (u["row"], u["col"])
        if key not in best or u["price"] < best[key]["price"]:
            best[key] = u
    return list(best.values())


def _save_local_excel(updates: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Синхронизация цен"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Конкурент", "Тип", "Модель парсера", "Модель Sheet", "Ремонт парсера",
               "Ремонт Sheet", "Качество", "Цена", "Ячейка Sheet"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for idx, u in enumerate(updates, 2):
        from gspread.utils import rowcol_to_a1
        cell_ref = rowcol_to_a1(u["row"], u["col"])
        ws.cell(row=idx, column=1, value=u["competitor"]).border = thin_border
        ws.cell(row=idx, column=2, value=u["device_type"]).border = thin_border
        ws.cell(row=idx, column=3, value=u["model"]).border = thin_border
        ws.cell(row=idx, column=4, value=u["sheet_model"]).border = thin_border
        ws.cell(row=idx, column=5, value=u["repair"]).border = thin_border
        ws.cell(row=idx, column=6, value=u["sheet_repair"]).border = thin_border
        ws.cell(row=idx, column=7, value=u["sheet_quality"]).border = thin_border
        ws.cell(row=idx, column=8, value=u["price"]).border = thin_border
        ws.cell(row=idx, column=9, value=cell_ref).border = thin_border

    for w, letter in [(18, "A"), (12, "B"), (22, "C"), (18, "D"), (25, "E"),
                      (25, "F"), (10, "G"), (10, "H"), (14, "I")]:
        ws.column_dimensions[letter].width = w

    wb.save(output_path)
    print(f"  Local Excel saved: {output_path}")


def sync_to_sheets(data: list[dict]):
    print(f"\n{'=' * 70}")
    print(f"  Syncing {len(data)} prices to Google Sheets...")
    print(f"{'=' * 70}")

    gc = gspread.service_account(CREDENTIALS_PATH)
    sh = gc.open_by_key(SPREADSHEET_ID)
    ws = sh.worksheet(SHEET_NAME)

    print("  Loading sheet index...")
    index, a_number_index = _load_sheet_index(ws)
    print(f"  Index built: {len(index)} rows, {len(a_number_index)} A-number entries")

    all_vals = ws.get_all_values()
    sheet_repairs = {"iPhone": [], "Macbook": []}
    for rv in all_vals:
        g = rv[6].strip() if len(rv) > 6 else ""
        j = rv[9].strip() if len(rv) > 9 else ""
        if g == "iPhone" and j and j not in sheet_repairs["iPhone"]:
            sheet_repairs["iPhone"].append(j)
        elif g == "Macbook" and j and j not in sheet_repairs["Macbook"]:
            sheet_repairs["Macbook"].append(j)

    iphone_data = [d for d in data if d.get("device_type") in ("iphone", "iPhone")]
    macbook_data = [d for d in data if d.get("device_type") in ("macbook", "Macbook")]

    iphone_updates = _build_updates(
        iphone_data, index, sheet_repairs,
        IPHONE_MODEL_TO_SHEET, IPHONE_REPAIR_TO_SHEET,
        a_number_index,
    )
    macbook_updates = _build_updates(
        macbook_data, index, sheet_repairs,
        MACBOOK_MODEL_TO_SHEET, MACBOOK_REPAIR_TO_SHEET,
        a_number_index,
    )

    all_updates = iphone_updates + macbook_updates
    all_updates = _deduplicate_updates(all_updates)

    print(f"  iPhone matches: {len(iphone_updates)}")
    print(f"  MacBook matches: {len(macbook_updates)}")
    print(f"  Total unique cells: {len(all_updates)}")

    if not all_updates:
        print("  No cells to update.")
        return []

    date_str = datetime.now().strftime("%d.%m")

    updated_competitors = set(u["competitor"] for u in all_updates)
    date_cells = []
    for comp_name, col in COMPETITOR_COLUMNS.items():
        if comp_name in updated_competitors:
            date_cells.append(gspread.Cell(6, col, date_str))
    if date_cells:
        try:
            ws.update_cells(date_cells, value_input_option="USER_ENTERED")
            print(f"  Updated dates in row 6 for {len(date_cells)} competitors: {date_str}")
        except Exception as e:
            print(f"  [ERROR] date row update: {e}")

    batch_size = 100
    for i in range(0, len(all_updates), batch_size):
        batch = all_updates[i:i + batch_size]
        cells = [gspread.Cell(u["row"], u["col"], str(u["price"])) for u in batch]
        try:
            ws.update_cells(cells, value_input_option="USER_ENTERED")
            print(f"  Written batch {i // batch_size + 1}: {len(batch)} cells")
        except Exception as e:
            print(f"  [ERROR] batch update: {e}, falling back to single writes...")
            for u in batch:
                try:
                    ws.update_cell(u["row"], u["col"], u["price"])
                except Exception as e2:
                    print(f"    [ERROR] cell ({u['row']},{u['col']}): {e2}")

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    local_path = os.path.join(os.path.dirname(__file__), f"sync_log_{ts}.xlsx")
    _save_local_excel(all_updates, local_path)

    print(f"\n  Sync complete: {len(all_updates)} cells updated in Google Sheets")
    return all_updates
