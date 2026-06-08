from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_report(data: list[dict], output_path: str):
    wb = Workbook()

    _create_iphone_sheet(wb, data)
    _create_macbook_sheet(wb, data)
    _create_raw_sheet(wb, data)

    wb.save(output_path)
    print(f"\nReport saved: {output_path}")


def _create_iphone_sheet(wb: Workbook, data: list[dict]):
    from config import IPHONE_MODELS, REPAIR_TYPES_IPHONE, COMPETITORS

    ws = wb.active
    ws.title = "iPhone цены"

    iphone_data = [d for d in data if d.get("device_type") == "iphone"]
    competitor_names = [c.name for c in COMPETITORS]

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    model_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1
    ws.cell(row=row, column=1, value=f"Срез цен конкурентов - iPhone (актуально на {datetime.now().strftime('%d.%m.%Y')})")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + len(competitor_names))
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)

    row = 3
    for repair_type in REPAIR_TYPES_IPHONE:
        ws.cell(row=row, column=1, value=repair_type)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + len(competitor_names))
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        row += 1
        ws.cell(row=row, column=1, value="Модель").font = header_font_white
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value="Min").font = header_font_white
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).border = thin_border
        for i, comp in enumerate(competitor_names, 3):
            ws.cell(row=row, column=i, value=comp).font = header_font_white
            ws.cell(row=row, column=i).fill = header_fill
            ws.cell(row=row, column=i).border = thin_border
            ws.cell(row=row, column=i).alignment = Alignment(horizontal="center", wrap_text=True)

        for model in IPHONE_MODELS:
            row += 1
            ws.cell(row=row, column=1, value=model).font = Font(bold=True, size=9)
            ws.cell(row=row, column=1).fill = model_fill
            ws.cell(row=row, column=1).border = thin_border

            prices_by_comp = {}
            for d in iphone_data:
                if d["model"] == model and d["repair"] == repair_type:
                    comp_name = d["competitor"]
                    p = d["price"]
                    if comp_name not in prices_by_comp or p < prices_by_comp[comp_name]:
                        prices_by_comp[comp_name] = p

            all_prices = list(prices_by_comp.values())
            min_price = min(all_prices) if all_prices else None

            ws.cell(row=row, column=2, value=min_price).border = thin_border
            ws.cell(row=row, column=2).font = Font(bold=True, size=9, color="008000")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

            for i, comp in enumerate(competitor_names, 3):
                val = prices_by_comp.get(comp)
                cell = ws.cell(row=row, column=i, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if val and min_price and val == min_price:
                    cell.font = Font(bold=True, size=9, color="008000")
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        row += 2

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8
    for i in range(3, 3 + len(competitor_names)):
        ws.column_dimensions[get_column_letter(i)].width = 13


def _create_macbook_sheet(wb: Workbook, data: list[dict]):
    from config import MACBOOK_MODELS, REPAIR_TYPES_MACBOOK, COMPETITORS

    ws = wb.create_sheet("MacBook цены")

    macbook_data = [d for d in data if d.get("device_type") == "macbook"]
    competitor_names = [c.name for c in COMPETITORS]

    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    model_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1
    ws.cell(row=row, column=1, value=f"Срез цен конкурентов - MacBook (актуально на {datetime.now().strftime('%d.%m.%Y')})")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + len(competitor_names))
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)

    row = 3
    for repair_type in REPAIR_TYPES_MACBOOK:
        ws.cell(row=row, column=1, value=repair_type)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + len(competitor_names))
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        row += 1
        ws.cell(row=row, column=1, value="Модель").font = header_font_white
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value="Min").font = header_font_white
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).border = thin_border
        for i, comp in enumerate(competitor_names, 3):
            ws.cell(row=row, column=i, value=comp).font = header_font_white
            ws.cell(row=row, column=i).fill = header_fill
            ws.cell(row=row, column=i).border = thin_border
            ws.cell(row=row, column=i).alignment = Alignment(horizontal="center", wrap_text=True)

        for model in MACBOOK_MODELS:
            row += 1
            ws.cell(row=row, column=1, value=model).font = Font(bold=True, size=9)
            ws.cell(row=row, column=1).fill = model_fill
            ws.cell(row=row, column=1).border = thin_border

            prices_by_comp = {}
            for d in macbook_data:
                if d["model"] == model and d["repair"] == repair_type:
                    comp_name = d["competitor"]
                    p = d["price"]
                    if comp_name not in prices_by_comp or p < prices_by_comp[comp_name]:
                        prices_by_comp[comp_name] = p

            all_prices = list(prices_by_comp.values())
            min_price = min(all_prices) if all_prices else None

            ws.cell(row=row, column=2, value=min_price).border = thin_border
            ws.cell(row=row, column=2).font = Font(bold=True, size=9, color="008000")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

            for i, comp in enumerate(competitor_names, 3):
                val = prices_by_comp.get(comp)
                cell = ws.cell(row=row, column=i, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if val and min_price and val == min_price:
                    cell.font = Font(bold=True, size=9, color="008000")
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        row += 2

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 8
    for i in range(3, 3 + len(competitor_names)):
        ws.column_dimensions[get_column_letter(i)].width = 13


def _create_raw_sheet(wb: Workbook, data: list[dict]):
    ws = wb.create_sheet("Сырые данные")

    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Конкурент", "Тип устройства", "Модель", "Тип ремонта", "Цена (руб)"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border

    for idx, d in enumerate(data, 2):
        ws.cell(row=idx, column=1, value=d.get("competitor", "")).border = thin_border
        ws.cell(row=idx, column=2, value=d.get("device_type", "")).border = thin_border
        ws.cell(row=idx, column=3, value=d.get("model", "")).border = thin_border
        ws.cell(row=idx, column=4, value=d.get("repair", "")).border = thin_border
        ws.cell(row=idx, column=5, value=d.get("price")).border = thin_border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 12
