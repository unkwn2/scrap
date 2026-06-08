import re
import csv
import io
import time
import requests
from bs4 import BeautifulSoup
from normalize import normalize_model, normalize_repair, extract_price

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
})
SESSION.verify = True


def fetch(url: str, timeout: int = 20) -> BeautifulSoup | None:
    try:
        resp = SESSION.get(url, timeout=timeout)
        resp.encoding = resp.apparent_encoding or "utf-8"
        return BeautifulSoup(resp.text, "lxml")
    except Exception as e:
        print(f"  [ERROR] fetch {url}: {e}")
        return None


def _get_target_models():
    from config import IPHONE_MODELS
    return IPHONE_MODELS


def _collect_model_links(soup, base_url: str, href_pattern: str) -> dict[str, str]:
    model_pages = {}
    links = [a for a in soup.find_all('a', href=True) if href_pattern in a['href']]
    for a in links:
        href = a['href']
        link_text = a.get_text(strip=True)
        model = normalize_model(link_text)
        if model and href:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href
    return model_pages


def _parse_model_page_tables(page_url: str, model: str) -> list[dict]:
    psoup = fetch(page_url)
    if not psoup:
        return []
    results = []
    tables = psoup.find_all("table")
    for table in tables:
        rows = table.find_all("tr")
        for row in rows:
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                repair_text = cells[0].get_text(strip=True)
                price_text = cells[1].get_text(strip=True)
                repair = normalize_repair(repair_text)
                price = extract_price(price_text)
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                    })
    return results


def parse_cepco(url: str) -> list[dict]:
    soup = fetch(url)
    if not soup:
        return []
    results = []
    sections = soup.find_all("div", class_="order__tab-accordion-item-content")
    for section in sections:
        header_div = section.find_previous_sibling("div", class_="order__tab-header-row")
        if not header_div:
            header_div = section.parent.find("div", class_="order__tab-header-row")
        section_repair = None
        if header_div:
            header_text = header_div.get_text(strip=True)
            section_repair = normalize_repair(header_text)
        rows = section.find_all("div", class_="order__tab-row")
        for row in rows:
            cols = row.find_all("div", class_="order__tab-col")
            if len(cols) >= 2:
                model_text = cols[0].get_text(strip=True)
                price_text = cols[1].get_text(strip=True)
                model = normalize_model(model_text)
                price = extract_price(price_text)
                repair = normalize_repair(model_text) or section_repair
                if model and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair or "Неизвестно",
                        "price": price,
                    })
    return results


def parse_hardworkers(base_url: str, iphone_url: str) -> list[dict]:
    soup = fetch(iphone_url)
    if not soup:
        return []
    results = []
    model_pages = _collect_model_links(soup, base_url, "remont-iphone")
    if not model_pages:
        model_pages = _collect_model_links(soup, base_url, "/iphone")
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        price_rows = psoup.find_all("div", class_="prices__table-row")
        if not price_rows:
            price_rows = psoup.find_all("div", class_="prices-damage__table-row")
        for row in price_rows:
            name_div = row.find("div", class_="prices__table-name")
            price_div = row.find("div", class_="prices__table-price")
            if not name_div or not price_div:
                name_div = row.find("div", class_="prices-damage__table-name")
                price_div = row.find("div", class_="prices-damage__table-price")
            if name_div and price_div:
                name_text = name_div.get_text(strip=True)
                price_text = price_div.get_text(strip=True)
                repair = normalize_repair(name_text)
                price = extract_price(price_text)
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                    })
        if not price_rows:
            results.extend(_parse_model_page_tables(page_url, model))
    return results


def parse_applepro(base_url: str, iphone_url: str) -> list[dict]:
    soup = fetch(iphone_url)
    if not soup:
        return []
    results = []
    model_pages = _collect_model_links(soup, base_url, "remont-iphone/iphone-")
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        rows = psoup.find_all("tr")
        for row in rows:
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                repair_text = cells[0].get_text(strip=True)
                price_text = cells[1].get_text(strip=True)
                repair = normalize_repair(repair_text)
                price = extract_price(price_text)
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                    })
        service_divs = psoup.find_all("div", class_="service-item")
        for div in service_divs:
            name_el = div.find("div", class_="service-name") or div.find("a")
            price_el = div.find("div", class_="service-price") or div.find("span", class_="price")
            if name_el and price_el:
                repair = normalize_repair(name_el.get_text(strip=True))
                price = extract_price(price_el.get_text(strip=True))
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                    })
    return results


def parse_modmac(base_url: str) -> list[dict]:
    soup = fetch(base_url)
    if not soup:
        return []
    results = []
    all_links = [a for a in soup.find_all('a', href=True) if '/services/iphone/remont-iphone-' in a['href']]
    model_pages = {}
    for a in all_links:
        href = a['href']
        last_segment = href.rstrip('/').split('/')[-1]
        if 'zamena' in last_segment or 'zamena-' in last_segment:
            continue
        link_text = a.get_text(strip=True)
        model = normalize_model(link_text)
        if model and href:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        price_blocks = psoup.find_all("div", class_="services_list_price_block")
        for pb in price_blocks:
            row = pb.parent.parent if pb.parent else None
            if not row:
                continue
            info = row.find("div", class_="services_list_info")
            if info:
                a_tag = info.find("a")
                name = a_tag.get_text(strip=True) if a_tag else info.get_text(strip=True)
                name = re.sub(r'\s+', ' ', name).strip()
            else:
                continue
            price_new = pb.find("div", class_="services_list_price_new")
            price_el = price_new if price_new else pb.find("div", class_="services_list_price")
            price_text = price_el.get_text(strip=True) if price_el else ""
            repair = normalize_repair(name)
            price = extract_price(price_text)
            if repair and price and price > 100:
                results.append({
                    "model": model,
                    "repair": repair,
                    "price": price,
                })
        results.extend(_parse_model_page_tables(page_url, model))
    return results


def parse_dabro(base_url: str) -> list[dict]:
    soup = fetch(base_url)
    if not soup:
        return []
    results = []
    model_pages = _collect_model_links(soup, base_url, "remont-iphone-")
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        tables = psoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    repair_text = cells[0].get_text(strip=True)
                    repair = normalize_repair(repair_text)
                    if not repair:
                        continue
                    price_div = cells[1].find("div", class_="price-time")
                    if price_div:
                        price_line = price_div.get_text(strip=True)
                        price_line = price_line.split("Ремонт")[0].strip()
                        price_line = price_line.split("Заказать")[0].strip()
                    else:
                        price_line = cells[1].get_text(strip=True)
                    price = extract_price(price_line)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                        })
    return results


def parse_planetiphone(base_url: str) -> list[dict]:
    soup = fetch(base_url)
    if not soup:
        return []
    results = []
    all_links = [a for a in soup.find_all('a', href=True) if 'remont-iphone-' in a['href'] and '.html' in a['href']]
    model_pages = {}
    for a in all_links:
        href = a['href']
        link_text = a.get_text(strip=True)
        model = normalize_model(link_text)
        if model and href:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        tables = psoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    repair_text = cells[0].get_text(strip=True)
                    price_cell = cells[1]
                    bold = price_cell.find("b")
                    if bold:
                        price_text = bold.get_text(strip=True)
                    else:
                        price_text = price_cell.get_text(strip=True)
                    repair = normalize_repair(repair_text)
                    if not repair:
                        continue
                    price = extract_price(price_text)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                        })
    return results


DISPLEYMASTER_URL_PATTERNS = [
    ("steklo", "Замена стекла"),
    ("disp", "Замена дисплея"),
    ("zadsteklo", "Замена заднего стекла"),
    ("akkum", "Замена аккумулятора"),
    ("batter", "Замена аккумулятора"),
]

DISPLEYMASTER_EXTRA_PAGES = [
    "https://displeymaster.ru/battery/",
    "https://displeymaster.ru/zamena-matricy-macbook/",
]


def parse_displeymaster(base_url: str) -> list[dict]:
    soup = fetch(base_url)
    if not soup:
        return []
    results = []
    iphone_links = [a for a in soup.find_all('a', href=True)
                    if any(p[0] in a['href'].lower() for p in DISPLEYMASTER_URL_PATTERNS)
                    and 'iphone' in a.get_text(strip=True).lower()]
    pages = {}
    for a in iphone_links:
        href = a['href']
        text = a.get_text(strip=True)
        model = normalize_model(text)
        if not model:
            continue
        repair = None
        href_lower = href.lower()
        for pattern, repair_name in DISPLEYMASTER_URL_PATTERNS:
            if pattern in href_lower:
                repair = repair_name
                break
        if not repair:
            repair = normalize_repair(text)
        if not repair:
            continue
        if not href.startswith("http"):
            href = base_url.rstrip("/") + "/" + href.lstrip("/")
        key = (model, repair)
        if key not in pages:
            pages[key] = (href, repair)

    target = _get_target_models()
    for (model, repair_type), (page_url, _) in pages.items():
        if model not in target:
            continue
        time.sleep(0.2)
        psoup = fetch(page_url)
        if not psoup:
            continue
        tables = psoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) < 2:
                    continue
                row_text = cells[0].get_text(strip=True)
                if not row_text:
                    continue
                row_repair = normalize_repair(row_text) or repair_type
                price = None
                for cell in reversed(cells[1:]):
                    p = extract_price(cell.get_text(strip=True))
                    if p and p > 100:
                        price = p
                        break
                if price:
                    results.append({
                        "model": model,
                        "repair": row_repair,
                        "price": price,
                    })

    for extra_url in DISPLEYMASTER_EXTRA_PAGES:
        time.sleep(0.2)
        esoup = fetch(extra_url)
        if not esoup:
            continue
        is_macbook = "macbook" in extra_url.lower()
        tables = esoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) < 2:
                    continue
                row_text = cells[0].get_text(strip=True)
                if not row_text:
                    continue
                if is_macbook:
                    from normalize import normalize_model_macbook
                    model = normalize_model_macbook(row_text)
                    if not model:
                        continue
                    repair = "Замена дисплея (матрицы)"
                    device_type = "macbook"
                else:
                    model = normalize_model(row_text)
                    if not model:
                        continue
                    repair = "Замена аккумулятора"
                    device_type = "iphone"
                price = None
                for cell in reversed(cells[1:]):
                    p = extract_price(cell.get_text(strip=True))
                    if p and p > 100:
                        price = p
                        break
                if price:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                        "device_type": device_type,
                    })
    return results


def parse_fixedone(base_url: str) -> list[dict]:
    results = []
    for url in ["https://service.fixed.one/iphonerepair/", "https://service.fixed.one/price/"]:
        soup = fetch(url)
        if not soup:
            continue
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            if not rows:
                continue
            header_row = rows[0]
            headers = []
            for cell in header_row.find_all(["th", "td"]):
                headers.append(cell.get_text(strip=True).lower())
            if not headers:
                continue
            for row in rows[1:]:
                cells = row.find_all(["td", "th"])
                if not cells:
                    continue
                model_text = cells[0].get_text(strip=True)
                model = normalize_model(model_text)
                if not model:
                    continue
                for i in range(1, min(len(cells), len(headers))):
                    repair = normalize_repair(headers[i])
                    if not repair:
                        continue
                    cell_text = cells[i].get_text(strip=True)
                    if "по запросу" in cell_text.lower() or "—" in cell_text:
                        continue
                    price = extract_price(cell_text)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                        })
    return results


def parse_kibercentre(base_url: str) -> list[dict]:
    soup = fetch(base_url)
    if not soup:
        return []
    results = []
    all_links = [a for a in soup.find_all('a', href=True)
                 if 'remont_iphone_' in a['href'] or 'iphone_' in a['href']]
    model_pages = {}
    for a in all_links:
        href = a['href']
        text = a.get_text(strip=True)
        model = normalize_model(text)
        if not model:
            href_segment = href.rstrip('/').split('/')[-1]
            segment_clean = re.sub(r'remont_iphone_', '', href_segment)
            segment_clean = re.sub(r'iphone_', '', segment_clean)
            segment_clean = segment_clean.replace('_', ' ')
            model = normalize_model(segment_clean)
        if model and href:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        tables = psoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    repair_text = cells[0].get_text(strip=True)
                    price_text = cells[1].get_text(strip=True)
                    repair = normalize_repair(repair_text)
                    if not repair:
                        continue
                    price = extract_price(price_text)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                        })
        if not tables:
            price_containers = psoup.find_all("div", class_="iphonebanner__pricecont")
            for pc in price_containers:
                label = pc.find_previous("div", class_="iphonebanner__label")
                if not label:
                    label = pc.parent.find_previous("div", class_="iphonebanner__label")
                repair = normalize_repair(label.get_text(strip=True)) if label else None
                price_val = pc.find("div", class_="iphonebanner__priceval")
                if not price_val:
                    price_val = pc.find("div", class_="iphonebanner__price")
                if price_val:
                    price = extract_price(price_val.get_text(strip=True))
                    if repair and price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                        })
    return results


def parse_google_sheets(url: str) -> list[dict]:
    gid_match = re.search(r"gid=(\d+)", url)
    base = url.split("/pubhtml")[0].split("/pub?")[0].split("/pub")[0]
    if gid_match:
        csv_url = f"{base}/pub?gid={gid_match.group(1)}&single=true&output=csv"
    else:
        csv_url = f"{base}/pub?output=csv"
    try:
        resp = SESSION.get(csv_url, timeout=30)
        resp.encoding = resp.apparent_encoding or "utf-8"
        text = resp.text
        if len(text) < 20:
            print(f"  [WARN] google sheets: empty response ({len(text)} bytes)")
            return []
    except Exception as e:
        print(f"  [ERROR] google sheets: {e}")
        return []

    results = []
    reader = csv.reader(io.StringIO(text))
    header_row = None
    header_idx = -1
    for i, row in enumerate(reader):
        if not row:
            continue
        first = row[0].strip().lower() if row else ""
        if first in ["модель", "model"] or ("замена" in " ".join(row).lower() and header_row is None):
            header_row = row
            header_idx = i
            continue
        if header_row is None:
            continue
        model_text = row[0].strip() if row else ""
        if not model_text:
            continue
        model = normalize_model(model_text)
        if not model:
            continue
        for j in range(1, min(len(row), len(header_row))):
            header = header_row[j].strip() if j < len(header_row) else ""
            if not header:
                continue
            repair = normalize_repair(header)
            if not repair:
                continue
            cell = row[j].strip() if j < len(row) else ""
            prices = cell.split("/")
            for p in prices:
                price = extract_price(p.strip())
                if price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                    })
                    break
    return results

JABUKA_MODEL_ORDER = [
    "iPhone 17", "iPhone 17 Air", "iPhone 17 Pro", "iPhone 17 Pro Max",
    "iPhone 16", "iPhone 16 Plus", "iPhone 16 Pro", "iPhone 16 Pro Max",
    "iPhone 15", "iPhone 15 Plus", "iPhone 15 Pro", "iPhone 15 Pro Max",
    "iPhone 14", "iPhone 14 Plus", "iPhone 14 Pro", "iPhone 14 Pro Max",
    "iPhone 13", "iPhone 13 Pro", "iPhone 13 Pro Max", "iPhone 13 mini",
    "iPhone 12", "iPhone 12 Pro", "iPhone 12 Pro Max", "iPhone 12 mini",
    "iPhone 11", "iPhone 11 Pro", "iPhone 11 Pro Max",
    "iPhone X", "iPhone XS", "iPhone XS Max", "iPhone XR",
    "iPhone 8", "iPhone 8 Plus", "iPhone 7", "iPhone 7 Plus",
    "iPhone 6", "iPhone 6s", "iPhone 6 Plus",
]


def parse_jabuka(url: str) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [INFO] playwright not available for jabuka")
        return []

    results = []
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, timeout=25000)
            page.wait_for_timeout(5000)
            content = page.content()
            browser.close()

        soup = BeautifulSoup(content, "lxml")
        recs = soup.find_all("div", class_="t-rec", attrs={"data-record-type": "396"})

        model_list = [m for m in JABUKA_MODEL_ORDER if m in _get_target_models()]

        price_blocks = []
        for rec in recs:
            text = rec.get_text(separator=" | ", strip=True)
            if "замена" in text.lower() and "₽" in text:
                price_blocks.append(text)

        for i, block_text in enumerate(price_blocks):
            if i >= len(model_list):
                break
            model = model_list[i]
            pairs = re.findall(r'(Замена[^₽]*?)\s*(\d[\d\s]*)₽', block_text)
            for repair_text, price_text in pairs:
                repair = normalize_repair(repair_text)
                price = extract_price(price_text)
                if repair and price and price > 100:
                    results.append({"model": model, "repair": repair, "price": price})
    except Exception as e:
        print(f"  [ERROR] jabuka: {e}")
    return results


def parse_isupport(index_url: str) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [INFO] playwright not available for isupport")
        return []

    results = []
    base = "https://www.isupport.ru"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            page.goto(index_url.rstrip("/") + "/", timeout=60000)
            page.wait_for_timeout(4000)

            links = page.query_selector_all("a")
            model_pages = {}
            for a in links:
                href = a.get_attribute("href") or ""
                text = (a.inner_text() or "").strip()
                model = normalize_model(text)
                if model and href and "/repair/repair-iphone/iphone-" in href:
                    if model not in model_pages:
                        full = href if href.startswith("http") else base + href
                        model_pages[model] = full

            target = _get_target_models()
            for model, page_url in sorted(model_pages.items()):
                if model not in target:
                    continue
                try:
                    page.goto(page_url, timeout=30000)
                except Exception:
                    continue
                page.wait_for_timeout(3000)
                try:
                    for _ in range(8):
                        page.evaluate("window.scrollBy(0, 600)")
                        page.wait_for_timeout(200)
                except Exception:
                    pass
                page.wait_for_timeout(1000)

                content = page.content()
                soup = BeautifulSoup(content, "lxml")
                items = soup.find_all("div", class_="service-item")
                for item in items:
                    try:
                        title_el = item.find("div", class_="service-item-title")
                        price_el = item.find("div", class_="service-item-price")
                        if not title_el or not price_el:
                            continue
                        title = title_el.get_text(strip=True)
                        price_text = price_el.get_text(strip=True)
                        repair = normalize_repair(title)
                        price = extract_price(price_text)
                        if repair and price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                            })
                    except Exception:
                        pass

            browser.close()
    except Exception as e:
        print(f"  [ERROR] isupport playwright: {e}")
    return results


MOSDISPLAY_FILE_PATH = r"C:\devin_ai_powershell_devin\competitor_scraper_irepair\ЦЕНЫ Mosdisplay"


def _load_mosdisplay_workbook():
    import os
    if os.path.exists(MOSDISPLAY_FILE_PATH + ".xlsx"):
        import openpyxl
        return openpyxl.load_workbook(MOSDISPLAY_FILE_PATH + ".xlsx", data_only=True)
    if os.path.exists(MOSDISPLAY_FILE_PATH + ".xls"):
        try:
            import openpyxl
            return openpyxl.load_workbook(MOSDISPLAY_FILE_PATH + ".xls", data_only=True)
        except Exception:
            pass
        try:
            import xlrd
            return xlrd.open_workbook(MOSDISPLAY_FILE_PATH + ".xls")
        except Exception:
            pass
    return None


def _iter_ws_rows(ws, is_xlrd=False):
    if is_xlrd:
        for rx in range(ws.nrows):
            yield [ws.cell_value(rx, cx) for cx in range(ws.ncols)]
    else:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            yield [str(c.value or "").strip() for c in row]


def parse_mosdisplay(_) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("  [INFO] openpyxl not available for mosdisplay")
        return []

    results = []
    try:
        wb = _load_mosdisplay_workbook()
        if wb is None:
            print("  [WARN] mosdisplay: file not found (ЦЕНЫ Mosdisplay.xlsx/.xls)")
            return []
        is_xlrd = hasattr(wb, 'sheet_by_index')
        sheet_names = wb.sheet_names() if is_xlrd else wb.sheetnames

        for si, sheet_name in enumerate(sheet_names):
            sn_lower = sheet_name.lower().strip()
            is_iphone = "iphone" in sn_lower or "\U0001f4f1" in sheet_name
            is_macbook = "macbook" in sn_lower
            if not is_iphone and not is_macbook:
                continue
            ws = wb.sheet_by_index(si) if is_xlrd else wb[sheet_name]

            if is_iphone:
                col_model_map = {}
                for vals in _iter_ws_rows(ws, is_xlrd):
                    series_text = str(vals[0]).strip() if vals else ""
                    if series_text and ("серия" in series_text.lower() or (
                        "iphone" in series_text.lower() and "ремонт" not in series_text.lower()
                    )):
                        col_model_map = {}
                        for i in range(8, min(len(vals), 18)):
                            cell_val = str(vals[i]).strip() if i < len(vals) else ""
                            if cell_val:
                                m = normalize_model(cell_val)
                                if m:
                                    col_model_map[i] = m
                        continue
                    repair_text = str(vals[2]).strip() if len(vals) > 2 else ""
                    if not repair_text:
                        continue
                    repair = normalize_repair(repair_text)
                    if not repair:
                        continue
                    for col_idx, model in col_model_map.items():
                        cell_val = str(vals[col_idx]).strip() if col_idx < len(vals) else ""
                        if not cell_val or cell_val in ["-", "", "Уточнять", "Нужно уточнить", "Недоступно"]:
                            continue
                        first_price = cell_val.split("/")[0].split("вместе")[0].split("отдельно")[0].strip()
                        first_price = re.split(r'\s', first_price)[0].strip()
                        price = extract_price(first_price)
                        if price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                                "device_type": "iphone",
                            })

            elif is_macbook:
                col_repair_map = {}
                for vals in _iter_ws_rows(ws, is_xlrd):
                    series_text = str(vals[0]).strip() if vals else ""
                    if series_text and ("серия" in series_text.lower() or "pro" in series_text.lower() or "air" in series_text.lower()):
                        col_repair_map = {}
                        for i in range(4, min(len(vals), 20)):
                            cell_val = str(vals[i]).strip() if i < len(vals) else ""
                            if cell_val:
                                repair = normalize_repair(cell_val)
                                if repair:
                                    col_repair_map[i] = repair
                        continue
                    if not col_repair_map:
                        continue
                    model_text = str(vals[2]).strip() if len(vals) > 2 else ""
                    marking = str(vals[3]).strip() if len(vals) > 3 else ""
                    from normalize import normalize_model_macbook
                    model = normalize_model_macbook(marking) or normalize_model_macbook(model_text)
                    if not model:
                        continue
                    for col_idx, repair in col_repair_map.items():
                        cell_val = str(vals[col_idx]).strip() if col_idx < len(vals) else ""
                        if not cell_val or cell_val in ["-", "", "Уточнять", "Нужно уточнить", "Недоступно"]:
                            continue
                        first_price = cell_val.split("/")[0].split("вместе")[0].split("отдельно")[0].strip()
                        first_price = re.split(r'\s', first_price)[0].strip()
                        price = extract_price(first_price)
                        if price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                                "device_type": "macbook",
                            })
        if not is_xlrd:
            wb.close()
    except Exception as e:
        print(f"  [ERROR] mosdisplay xlsx: {e}")
    return results


def parse_brobrolab(index_url: str) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [INFO] playwright not available, trying selenium for brobrolab")
        return _parse_brobrolab_selenium(index_url)

    results = []
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(index_url, timeout=20000)
            page.wait_for_timeout(3000)

            links = page.query_selector_all("a")
            model_pages = {}
            for a in links:
                href = a.get_attribute("href") or ""
                text = (a.inner_text() or "").strip().replace("\n", " ")
                model = normalize_model(text)
                if model and href and "brobrolab.ru/services-iphone" in href:
                    if model not in model_pages:
                        model_pages[model] = href

            target = _get_target_models()
            count = 0
            for model, page_url in sorted(model_pages.items()):
                if model not in target:
                    continue
                if count >= 20:
                    break
                count += 1
                try:
                    page.goto(page_url, timeout=20000)
                except Exception:
                    continue
                page.wait_for_timeout(2000)
                try:
                    for _ in range(6):
                        page.evaluate("window.scrollBy(0, 500)")
                        page.wait_for_timeout(200)
                except Exception:
                    pass
                page.wait_for_timeout(1000)

                try:
                    cards = page.query_selector_all(".t-store__card")
                except Exception:
                    continue
                for card in cards:
                    try:
                        title_el = card.query_selector(".js-store-prod-name, .t-store__card__title")
                        price_el = card.query_selector(".js-product-price, .js-store-prod-price-val")
                        if not title_el or not price_el:
                            continue
                        title = (title_el.inner_text() or "").strip()
                        price_text = (price_el.inner_text() or "").strip()
                        repair = normalize_repair(title)
                        price = extract_price(price_text)
                        if repair and price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                            })
                    except Exception:
                        pass

            browser.close()
    except Exception as e:
        print(f"  [ERROR] brobrolab playwright: {e}")
    return results


def _parse_brobrolab_selenium(index_url: str) -> list[dict]:
    try:
        from selenium import webdriver
        from selenium.webdriver.edge.options import Options
        from selenium.webdriver.common.by import By
    except ImportError:
        print("  [INFO] selenium not available either, skipping brobrolab")
        return []

    results = []
    driver = None
    try:
        options = Options()
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--window-size=1920,1080")
        driver = webdriver.Edge(options=options)
        driver.set_page_load_timeout(20)
        driver.get(index_url)
        time.sleep(5)

        links = driver.find_elements(By.TAG_NAME, "a")
        model_pages = {}
        for a in links:
            href = a.get_attribute('href') or ''
            text = a.text.strip().replace('\n', ' ')
            model = normalize_model(text)
            if model and href and 'brobrolab.ru/services-iphone' in href:
                if model not in model_pages:
                    model_pages[model] = href

        target = _get_target_models()
        count = 0
        for model, page_url in sorted(model_pages.items()):
            if model not in target:
                continue
            if count >= 20:
                break
            count += 1
            try:
                driver.get(page_url)
            except Exception:
                continue
            time.sleep(3)
            try:
                for _ in range(6):
                    driver.execute_script("window.scrollBy(0, 500);")
                    time.sleep(0.3)
            except Exception:
                pass
            time.sleep(2)
            try:
                cards = driver.find_elements(By.CSS_SELECTOR, ".t-store__card")
            except Exception:
                continue
            for card in cards:
                try:
                    title_el = card.find_element(By.CSS_SELECTOR, ".js-store-prod-name, .t-store__card__title")
                    price_el = card.find_element(By.CSS_SELECTOR, ".js-product-price, .js-store-prod-price-val")
                    title = title_el.text.strip()
                    price_text = price_el.text.strip()
                    repair = normalize_repair(title)
                    price = extract_price(price_text)
                    if repair and price and price > 100:
                        results.append({"model": model, "repair": repair, "price": price})
                except Exception:
                    pass
    except Exception as e:
        print(f"  [ERROR] brobrolab selenium: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass
    return results


def _parse_with_selenium(url: str) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
        return _parse_with_playwright(url)
    except ImportError:
        pass

    try:
        from selenium import webdriver
        from selenium.webdriver.edge.options import Options
        from selenium.webdriver.common.by import By
    except ImportError:
        print("  [INFO] no browser automation available")
        return []

    results = []
    driver = None
    try:
        options = Options()
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--window-size=1920,1080")
        driver = webdriver.Edge(options=options)
        driver.set_page_load_timeout(20)
        driver.get(url)
        time.sleep(4)
        for _ in range(6):
            driver.execute_script("window.scrollBy(0, 500);")
            time.sleep(0.3)
        time.sleep(2)
        page_source = driver.page_source
        driver.quit()
        driver = None

        soup = BeautifulSoup(page_source, "lxml")
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    repair_text = cells[0].get_text(strip=True)
                    price_text = cells[1].get_text(strip=True)
                    model = None
                    for m in _get_target_models():
                        if m.lower() in url.lower():
                            model = m
                            break
                    if not model:
                        model = normalize_model(repair_text) or normalize_model(url)
                    repair = normalize_repair(repair_text)
                    price = extract_price(price_text)
                    if repair and price and price > 100:
                        results.append({"model": model or "Неизвестно", "repair": repair, "price": price})
        if not results:
            all_els = soup.find_all(["div", "span", "p", "li"])
            for el in all_els:
                text = el.get_text(separator=" ", strip=True)
                model = normalize_model(text)
                if model:
                    price = extract_price(text)
                    repair = normalize_repair(text)
                    if price and price > 100:
                        results.append({"model": model, "repair": repair or "Неизвестно", "price": price})
    except Exception as e:
        print(f"  [ERROR] selenium: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass
    return results


def _parse_with_playwright(url: str) -> list[dict]:
    results = []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, timeout=20000)
            page.wait_for_timeout(4000)
            for _ in range(6):
                page.evaluate("window.scrollBy(0, 500)")
                page.wait_for_timeout(300)
            page.wait_for_timeout(2000)

            content = page.content()
            browser.close()

        soup = BeautifulSoup(content, "lxml")
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    repair_text = cells[0].get_text(strip=True)
                    price_text = cells[1].get_text(strip=True)
                    model = None
                    for m in _get_target_models():
                        if m.lower() in url.lower():
                            model = m
                            break
                    if not model:
                        model = normalize_model(repair_text) or normalize_model(url)
                    repair = normalize_repair(repair_text)
                    price = extract_price(price_text)
                    if repair and price and price > 100:
                        results.append({"model": model or "Неизвестно", "repair": repair, "price": price})
        if not results:
            all_els = soup.find_all(["div", "span", "p", "li"])
            for el in all_els:
                text = el.get_text(separator=" ", strip=True)
                model = normalize_model(text)
                if model:
                    price = extract_price(text)
                    repair = normalize_repair(text)
                    if price and price > 100:
                        results.append({"model": model, "repair": repair or "Неизвестно", "price": price})
    except Exception as e:
        print(f"  [ERROR] playwright: {e}")
    return results


APPLEPIE_SLUG_MAP = {
    "15": "iPhone 15", "15pro": "iPhone 15 Pro", "15promax": "iPhone 15 Pro Max",
    "15plus": "iPhone 15 Plus", "14": "iPhone 14", "14pro": "iPhone 14 Pro",
    "14promax": "iPhone 14 Pro Max", "14plus": "iPhone 14 Plus",
    "13": "iPhone 13", "13pro": "iPhone 13 Pro", "13promax": "iPhone 13 Pro Max",
    "13mini": "iPhone 13 mini", "12": "iPhone 12", "12pro": "iPhone 12 Pro",
    "12promax": "iPhone 12 Pro Max", "12mini": "iPhone 12 mini",
    "11promax": "iPhone 11 Pro Max", "11pro": "iPhone 11 Pro", "11": "iPhone 11",
    "XSmax": "iPhone XS Max", "XS": "iPhone XS", "X": "iPhone X", "XR": "iPhone XR",
}


def parse_applepie(base_url: str) -> list[dict]:
    soup = fetch(base_url)
    if not soup:
        return []
    results = []
    target = _get_target_models()

    price_links = [a for a in soup.find_all('a', href=True) if '/price-' in a['href']]
    model_pages = {}
    for a in price_links:
        href = a['href']
        slug = href.split('/price-')[-1].rstrip('/').split('?')[0].split('#')[0]
        model = APPLEPIE_SLUG_MAP.get(slug)
        if not model:
            link_text = a.get_text(strip=True)
            model = normalize_model(link_text)
        if model and model in target:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href

    for model, page_url in model_pages.items():
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue

        slots = psoup.find_all("div", class_="ms-slot--param-slot")
        for slot in slots:
            first_blk = slot.find("div", class_="blk_text")
            if not first_blk:
                continue
            service_name = first_blk.get_text(strip=True)

            yellow_boxes = slot.find_all("div", class_="blk_box_self")
            price_text = None
            for yb in yellow_boxes:
                style = yb.get("style", "") or yb.get("bg", "")
                if "ffd832" not in style.lower() and "#ffd832" not in style.lower():
                    bg_attr = yb.get("bg", "")
                    if "ffd832" not in bg_attr.lower():
                        continue
                bd = yb.find("div", class_="blk-data")
                if not bd:
                    continue
                txt = bd.get_text(strip=True)
                if "руб" in txt.lower() or re.match(r'^[\d\s]+руб', txt):
                    price_text = txt
                    break

            if not price_text:
                for yb in yellow_boxes:
                    bd = yb.find("div", class_="blk-data")
                    if bd:
                        txt = bd.get_text(strip=True)
                        p = extract_price(txt)
                        if p and p > 500 and not any(
                            kw in txt.lower()
                            for kw in ["мин", "час", "день", "бесплатно", "уточнять"]
                        ):
                            price_text = txt
                            break

            repair = normalize_repair(service_name)
            price = extract_price(price_text) if price_text else None
            if repair and price and price > 100:
                results.append({
                    "model": model,
                    "repair": repair,
                    "price": price,
                })

    return results


def parse_generic(url: str, competitor_name: str) -> list[dict]:
    soup = fetch(url)
    if not soup:
        return []
    results = []

    model_pages = _collect_model_links(soup, url.split("/")[0] + "//" + url.split("/")[2], "remont-iphone")
    target = _get_target_models()
    if model_pages:
        for model, page_url in model_pages.items():
            if model not in target:
                continue
            time.sleep(0.3)
            results.extend(_parse_model_page_tables(page_url, model))
        if results:
            return results

    tables = soup.find_all("table")
    for table in tables:
        rows = table.find_all("tr")
        headers = []
        if rows:
            first_row = rows[0]
            for cell in first_row.find_all(["th", "td"]):
                headers.append(cell.get_text(strip=True).lower())
        for row in rows[1:] if headers else rows:
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                model_text = cells[0].get_text(strip=True)
                model = normalize_model(model_text)
                if not model:
                    continue
                if len(headers) > 1:
                    for i in range(1, min(len(cells), len(headers))):
                        repair = normalize_repair(headers[i])
                        price = extract_price(cells[i].get_text(strip=True))
                        if price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair or "Неизвестно",
                                "price": price,
                            })
                else:
                    price = extract_price(cells[1].get_text(strip=True))
                    repair = normalize_repair(model_text)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair or "Неизвестно",
                            "price": price,
                        })

    if not results:
        price_containers = soup.find_all(
            ["div", "li", "span", "p"],
            class_=re.compile(r"price|cost|service|repair|remont", re.I),
        )
        for container in price_containers:
            text = container.get_text(separator=" ", strip=True)
            model = normalize_model(text)
            if model:
                price = extract_price(text)
                repair = normalize_repair(text)
                if price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair or "Неизвестно",
                        "price": price,
                    })

    if not results:
        all_text_els = soup.find_all(["div", "li", "span", "p"])
        for el in all_text_els:
            text = el.get_text(separator=" ", strip=True)
            if "iphone" not in text.lower():
                continue
            model = normalize_model(text)
            if model:
                price_match = re.search(r"[\d\s]{4,}р", text)
                if price_match:
                    price = extract_price(price_match.group())
                    repair = normalize_repair(text)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair or "Неизвестно",
                            "price": price,
                        })
    return results
