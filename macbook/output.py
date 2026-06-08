from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_report(data: list[dict], output_path: str):
    wb = Workbook()

    _create_macbook_sheet(wb, data)
    _create_raw_sheet(wb, data)

    wb.save(output_path)
    print(f"\nReport saved: {output_path}")


def _create_macbook_sheet(wb: Workbook, data: list[dict]):
    from config import MACBOOK_MODELS, REPAIR_TYPES_MACBOOK, QUALITY_TYPES, COMPETITORS

    ws = wb.active
    ws.title = "MacBook цены"

    competitor_names = [c.name for c in COMPETITORS]

    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    model_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    quality_fill_aasp = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    quality_fill_oem = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    min_font = Font(bold=True, size=9, color="008000")
    min_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    row = 1
    ws.cell(row=row, column=1, value=f"Срез цен конкурентов - MacBook (актуально на {datetime.now().strftime('%d.%m.%Y')})")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + len(competitor_names))
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)

    row = 3
    for repair_type in REPAIR_TYPES_MACBOOK:
        for quality in QUALITY_TYPES:
            repair_data = [d for d in data if d.get("repair") == repair_type and d.get("quality") == quality]
            if not repair_data:
                continue

            label = f"{repair_type} ({quality})"
            ws.cell(row=row, column=1, value=label)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + len(competitor_names))
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

            row += 1
            headers = ["Модель", "Качество", "Min"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=row, column=i, value=h).font = header_font_white
                ws.cell(row=row, column=i).fill = header_fill
                ws.cell(row=row, column=i).border = thin_border
            for i, comp in enumerate(competitor_names, 4):
                ws.cell(row=row, column=i, value=comp).font = header_font_white
                ws.cell(row=row, column=i).fill = header_fill
                ws.cell(row=row, column=i).border = thin_border
                ws.cell(row=row, column=i).alignment = Alignment(horizontal="center", wrap_text=True)

            for model in MACBOOK_MODELS:
                row += 1
                ws.cell(row=row, column=1, value=model).font = Font(bold=True, size=9)
                ws.cell(row=row, column=1).fill = model_fill
                ws.cell(row=row, column=1).border = thin_border

                qfill = quality_fill_aasp if quality == "AASP" else quality_fill_oem
                ws.cell(row=row, column=2, value=quality).font = Font(size=9)
                ws.cell(row=row, column=2).fill = qfill
                ws.cell(row=row, column=2).border = thin_border
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

                prices_by_comp = {}
                for d in repair_data:
                    if d["model"] == model:
                        comp_name = d["competitor"]
                        p = d["price"]
                        if comp_name not in prices_by_comp or p < prices_by_comp[comp_name]:
                            prices_by_comp[comp_name] = p

                all_prices = list(prices_by_comp.values())
                min_price = min(all_prices) if all_prices else None

                ws.cell(row=row, column=3, value=min_price).border = thin_border
                ws.cell(row=row, column=3).font = min_font
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

                for i, comp in enumerate(competitor_names, 4):
                    val = prices_by_comp.get(comp)
                    cell = ws.cell(row=row, column=i, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if val and min_price and val == min_price:
                        cell.font = min_font
                        cell.fill = min_fill

            row += 2

        repair_no_quality = [d for d in data if d.get("repair") == repair_type and not d.get("quality")]
        if repair_no_quality:
            label = f"{repair_type} (без указания качества)"
            ws.cell(row=row, column=1, value=label)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3 + len(competitor_names))
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

            row += 1
            headers = ["Модель", "Качество", "Min"]
            for i, h in enumerate(headers, 1):
                ws.cell(row=row, column=i, value=h).font = header_font_white
                ws.cell(row=row, column=i).fill = header_fill
                ws.cell(row=row, column=i).border = thin_border
            for i, comp in enumerate(competitor_names, 4):
                ws.cell(row=row, column=i, value=comp).font = header_font_white
                ws.cell(row=row, column=i).fill = header_fill
                ws.cell(row=row, column=i).border = thin_border
                ws.cell(row=row, column=i).alignment = Alignment(horizontal="center", wrap_text=True)

            for model in MACBOOK_MODELS:
                row += 1
                ws.cell(row=row, column=1, value=model).font = Font(bold=True, size=9)
                ws.cell(row=row, column=1).fill = model_fill
                ws.cell(row=row, column=1).border = thin_border

                ws.cell(row=row, column=2, value="—").font = Font(size=9)
                ws.cell(row=row, column=2).border = thin_border
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

                prices_by_comp = {}
                for d in repair_no_quality:
                    if d["model"] == model:
                        comp_name = d["competitor"]
                        p = d["price"]
                        if comp_name not in prices_by_comp or p < prices_by_comp[comp_name]:
                            prices_by_comp[comp_name] = p

                all_prices = list(prices_by_comp.values())
                min_price = min(all_prices) if all_prices else None

                ws.cell(row=row, column=3, value=min_price).border = thin_border
                ws.cell(row=row, column=3).font = min_font
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

                for i, comp in enumerate(competitor_names, 4):
                    val = prices_by_comp.get(comp)
                    cell = ws.cell(row=row, column=i, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if val and min_price and val == min_price:
                        cell.font = min_font
                        cell.fill = min_fill

            row += 2

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    for i in range(4, 4 + len(competitor_names)):
        ws.column_dimensions[get_column_letter(i)].width = 13


def _create_raw_sheet(wb: Workbook, data: list[dict]):
    ws = wb.create_sheet("Сырые данные")

    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Конкурент", "Модель", "Тип ремонта", "Качество", "Цена (руб)"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = thin_border

    for idx, d in enumerate(data, 2):
        ws.cell(row=idx, column=1, value=d.get("competitor", "")).border = thin_border
        ws.cell(row=idx, column=2, value=d.get("model", "")).border = thin_border
        ws.cell(row=idx, column=3, value=d.get("repair", "")).border = thin_border
        ws.cell(row=idx, column=4, value=d.get("quality", "")).border = thin_border
        ws.cell(row=idx, column=5, value=d.get("price")).border = thin_border

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
