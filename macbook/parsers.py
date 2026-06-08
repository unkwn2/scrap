import re
import csv
import io
import time
import requests
from bs4 import BeautifulSoup
from normalize import normalize_model_macbook, normalize_model_macbook_multi, normalize_repair, normalize_quality, extract_price, normalize_with_context_multi

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
})


def fetch(url: str, timeout: int = 20) -> BeautifulSoup | None:
    try:
        resp = SESSION.get(url, timeout=timeout)
        resp.encoding = resp.apparent_encoding or "utf-8"
        return BeautifulSoup(resp.text, "lxml")
    except Exception as e:
        print(f"  [ERROR] fetch {url}: {e}")
        return None


def _get_target_models():
    from config import MACBOOK_MODELS
    return MACBOOK_MODELS


def _collect_macbook_model_links(soup, base_url: str, href_pattern: str) -> dict[str, str]:
    model_pages = {}
    links = [a for a in soup.find_all('a', href=True) if href_pattern in a['href']]
    for a in links:
        href = a['href']
        link_text = a.get_text(strip=True)
        model = normalize_model_macbook(link_text)
        if model and href:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href
    return model_pages


def _parse_model_page_tables(page_url: str, model: str) -> list[dict]:
    psoup = fetch(page_url)
    if not psoup:
        return []
    results = []
    tables = psoup.find_all("table")
    for table in tables:
        rows = table.find_all("tr")
        for row in rows:
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                repair_text = cells[0].get_text(strip=True)
                price_text = cells[1].get_text(strip=True)
                repair = normalize_repair(repair_text)
                quality = normalize_quality(repair_text)
                price = extract_price(price_text)
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                        "quality": quality,
                    })
    return results


def parse_cepco(url: str) -> list[dict]:
    soup = fetch(url)
    if not soup:
        return []
    results = []

    model_links = soup.find_all("a", class_="phone_list_li")
    model_pages = {}
    for a in model_links:
        href = a.get("href", "")
        link_text = a.get_text(strip=True)
        model = normalize_model_macbook(link_text)
        if model and href:
            if not href.startswith("http"):
                href = url.rstrip("/").split("/")[0] + "//" + url.split("/")[2] + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href

    if not model_pages:
        sub_links = soup.find_all("a", href=True)
        for a in sub_links:
            href = a["href"]
            text = a.get_text(strip=True)
            if "macbook" in href.lower() and "remont-macbook" in href:
                model = normalize_model_macbook(text)
                if model and href:
                    if not href.startswith("http"):
                        base = url.rstrip("/").split("/")[0] + "//" + url.split("/")[2]
                        href = base + "/" + href.lstrip("/")
                    if model not in model_pages:
                        model_pages[model] = href

    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        sections = psoup.find_all("div", class_="order__tab-accordion-item-content")
        for section in sections:
            header_div = section.find_previous_sibling("div", class_="order__tab-header-row")
            if not header_div:
                header_div = section.parent.find("div", class_="order__tab-header-row")
            section_repair = None
            if header_div:
                header_text = header_div.get_text(strip=True)
                section_repair = normalize_repair(header_text)
            rows = section.find_all("div", class_="order__tab-row")
            for row in rows:
                cols = row.find_all("div", class_="order__tab-col")
                if len(cols) >= 2:
                    model_text = cols[0].get_text(strip=True)
                    price_text = cols[1].get_text(strip=True)
                    repair = normalize_repair(model_text) or section_repair
                    quality = normalize_quality(model_text)
                    price = extract_price(price_text)
                    if repair and price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                            "quality": quality,
                        })

    return results


def parse_hardworkers(base_url: str, macbook_url: str) -> list[dict]:
    soup = fetch(macbook_url)
    if not soup:
        return []
    results = []
    model_pages = _collect_macbook_model_links(soup, base_url, "remont-macbook")
    if not model_pages:
        model_pages = _collect_macbook_model_links(soup, base_url, "macbook")
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        price_rows = psoup.find_all("div", class_="prices__table-row")
        if not price_rows:
            price_rows = psoup.find_all("div", class_="prices-damage__table-row")
        for row in price_rows:
            name_div = row.find("div", class_="prices__table-name")
            price_div = row.find("div", class_="prices__table-price")
            if not name_div or not price_div:
                name_div = row.find("div", class_="prices-damage__table-name")
                price_div = row.find("div", class_="prices-damage__table-price")
            if name_div and price_div:
                name_text = name_div.get_text(strip=True)
                price_text = price_div.get_text(strip=True)
                repair = normalize_repair(name_text)
                quality = normalize_quality(name_text)
                price = extract_price(price_text)
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                        "quality": quality,
                    })
        if not price_rows:
            results.extend(_parse_model_page_tables(page_url, model))
    return results


DISPLEYMASTER_EXTRA_PAGES = [
    "https://displeymaster.ru/zamena-matricy-macbook/",
]


def parse_displeymaster(base_url: str) -> list[dict]:
    results = []
    for extra_url in DISPLEYMASTER_EXTRA_PAGES:
        time.sleep(0.2)
        esoup = fetch(extra_url)
        if not esoup:
            continue
        tables = esoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) < 2:
                    continue
                row_text = cells[0].get_text(strip=True)
                if not row_text:
                    continue
                model = normalize_model_macbook(row_text)
                if not model:
                    continue
                repair = normalize_repair(row_text) or "Замена матрицы"
                quality = normalize_quality(row_text)
                price = None
                for cell in reversed(cells[1:]):
                    p = extract_price(cell.get_text(strip=True))
                    if p and p > 100:
                        price = p
                        break
                if price:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                        "quality": quality,
                    })
    return results


KIBERCENTRE_MACBOOK_REPAIR_PAGES = [
    ("https://kibercentre.ru/servis_apple/macbook_zamena_matrici/", "Замена матрицы"),
    ("https://kibercentre.ru/servis_apple/macbook_zamena_klaviaturi/", "Замена клавиатуры"),
    ("https://kibercentre.ru/servis_apple/macbook_zamena_akkumulyatora/", "Замена аккумулятора"),
    ("https://kibercentre.ru/servis_apple/macbook_zamena_tachpada/", "Замена трекпада"),
    ("https://kibercentre.ru/servis_apple/macbook_zamena_ssd_disca", "Замена SSD"),
    ("https://kibercentre.ru/servis_apple/macbook_zamena_dinamika/", "Замена динамика"),
    ("https://kibercentre.ru/servis_apple/macbook_razyem_zaryadki/", None),
    ("https://kibercentre.ru/servis_apple/macbook_zamena_kulera/", "Замена вентилятора"),
    ("https://kibercentre.ru/servis_apple/macbook_remont_posle_vodi/", "Ремонт после попадания воды"),
]


def parse_kibercentre(base_url: str, macbook_url: str) -> list[dict]:
    results = []
    for page_url, default_repair in KIBERCENTRE_MACBOOK_REPAIR_PAGES:
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        tables = psoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    model_text = cells[0].get_text(strip=True)
                    price_text = cells[1].get_text(strip=True)
                    model = normalize_model_macbook(model_text)
                    if not model:
                        data_device = row.get("data-device", "")
                        if data_device:
                            model = normalize_model_macbook(data_device)
                    if not model:
                        continue
                    repair = normalize_repair(model_text) or default_repair
                    quality = normalize_quality(model_text)
                    price = extract_price(price_text)
                    if price and price > 100 and repair:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                            "quality": quality,
                        })

    if not results:
        soup = fetch(macbook_url)
        if not soup:
            return []
        all_links = [a for a in soup.find_all('a', href=True)
                     if 'macbook' in a['href'].lower() and 'servis_apple' in a['href']]
        for a in all_links:
            href = a['href']
            text = a.get_text(strip=True)
            repair = normalize_repair(text)
            if not repair:
                continue
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            time.sleep(0.3)
            psoup = fetch(href)
            if not psoup:
                continue
            tables = psoup.find_all("table")
            for table in tables:
                rows = table.find_all("tr")
                for row in rows:
                    cells = row.find_all(["td", "th"])
                    if len(cells) >= 2:
                        model_text = cells[0].get_text(strip=True)
                        price_text = cells[1].get_text(strip=True)
                        model = normalize_model_macbook(model_text)
                        if not model:
                            data_device = row.get("data-device", "")
                            if data_device:
                                model = normalize_model_macbook(data_device)
                        if not model:
                            continue
                        quality = normalize_quality(model_text)
                        price = extract_price(price_text)
                        if price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                                "quality": quality,
                            })

    return results


def _normalize_with_context(model_text: str, section_title: str) -> str | None:
    models = normalize_with_context_multi(model_text, section_title)
    return models[0] if models else None


def parse_fixedone(base_url: str) -> list[dict]:
    results = []
    for url in ["https://service.fixed.one/mbprepair/", "https://service.fixed.one/mbarepair/"]:
        soup = fetch(url)
        if not soup:
            continue
        toggles = soup.find_all("h5", class_="et_pb_toggle_title")
        if not toggles:
            tables = soup.find_all("table")
            for table in tables:
                rows = table.find_all("tr")
                if not rows:
                    continue
                header_row = rows[0]
                headers = []
                for cell in header_row.find_all(["th", "td"]):
                    headers.append(cell.get_text(strip=True).lower())
                for row in rows[1:]:
                    cells = row.find_all(["td", "th"])
                    if not cells:
                        continue
                    model_text = cells[0].get_text(strip=True)
                    model = normalize_model_macbook(model_text)
                    if not model:
                        continue
                    for i in range(1, min(len(cells), len(headers))):
                        repair = normalize_repair(headers[i])
                        if not repair:
                            continue
                        cell_text = cells[i].get_text(strip=True)
                        if "по запросу" in cell_text.lower() or "—" in cell_text:
                            continue
                        price = extract_price(cell_text)
                        if price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                                "quality": "",
                            })
            continue

        for toggle in toggles:
            section_title = toggle.get_text(strip=True)
            content = toggle.find_next_sibling("div", class_="et_pb_toggle_content")
            if not content:
                continue
            table = content.find("table")
            if not table:
                continue
            rows = table.find_all("tr")
            if not rows:
                continue
            header_row = rows[0]
            headers = []
            for cell in header_row.find_all(["th", "td"]):
                headers.append(cell.get_text(strip=True).lower())
            for row in rows[1:]:
                cells = row.find_all(["td", "th"])
                if not cells:
                    continue
                model_text = cells[0].get_text(strip=True)
                models = normalize_with_context_multi(model_text, section_title)
                if not models:
                    continue
                for i in range(1, min(len(cells), len(headers))):
                    repair = normalize_repair(headers[i])
                    if not repair:
                        continue
                    cell_text = cells[i].get_text(strip=True)
                    if "по запросу" in cell_text.lower() or "—" in cell_text:
                        continue
                    price = extract_price(cell_text)
                    if price and price > 100:
                        for model in models:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                                "quality": "",
                            })
    return results


def parse_applepro(base_url: str, macbook_url: str) -> list[dict]:
    soup = fetch(macbook_url)
    if not soup:
        return []
    results = []
    model_pages = _collect_macbook_model_links(soup, base_url, "remont-macbook/macbook-")
    if not model_pages:
        model_pages = _collect_macbook_model_links(soup, base_url, "remont-macbook/")
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        rows = psoup.find_all("tr")
        for row in rows:
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                repair_text = cells[0].get_text(strip=True)
                price_text = cells[1].get_text(strip=True)
                repair = normalize_repair(repair_text)
                quality = normalize_quality(repair_text)
                price = extract_price(price_text)
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                        "quality": quality,
                    })
        service_divs = psoup.find_all("div", class_="service-item")
        for div in service_divs:
            name_el = div.find("div", class_="service-name") or div.find("a")
            price_el = div.find("div", class_="service-price") or div.find("span", class_="price")
            if name_el and price_el:
                repair = normalize_repair(name_el.get_text(strip=True))
                quality = normalize_quality(name_el.get_text(strip=True))
                price = extract_price(price_el.get_text(strip=True))
                if repair and price and price > 100:
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                        "quality": quality,
                    })
    return results


def parse_brobrolab(index_url: str) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [INFO] playwright not available for brobrolab")
        return []

    results = []
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(index_url, timeout=20000)
            page.wait_for_timeout(3000)

            links = page.query_selector_all("a")
            model_pages = {}
            for a in links:
                href = a.get_attribute("href") or ""
                text = (a.inner_text() or "").strip().replace("\n", " ")
                model = normalize_model_macbook(text)
                if model and href and "brobrolab.ru/services" in href:
                    if model not in model_pages:
                        model_pages[model] = href

            target = _get_target_models()
            count = 0
            for model, page_url in sorted(model_pages.items()):
                if model not in target:
                    continue
                if count >= 20:
                    break
                count += 1
                try:
                    page.goto(page_url, timeout=20000)
                except Exception:
                    continue
                page.wait_for_timeout(2000)
                try:
                    for _ in range(6):
                        page.evaluate("window.scrollBy(0, 500)")
                        page.wait_for_timeout(200)
                except Exception:
                    pass
                page.wait_for_timeout(1000)

                try:
                    cards = page.query_selector_all(".t-store__card")
                except Exception:
                    continue
                for card in cards:
                    try:
                        title_el = card.query_selector(".js-store-prod-name, .t-store__card__title")
                        price_el = card.query_selector(".js-product-price, .js-store-prod-price-val")
                        if not title_el or not price_el:
                            continue
                        title = (title_el.inner_text() or "").strip()
                        price_text = (price_el.inner_text() or "").strip()
                        repair = normalize_repair(title)
                        quality = normalize_quality(title)
                        price = extract_price(price_text)
                        if repair and price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                                "quality": quality,
                            })
                    except Exception:
                        pass

            browser.close()
    except Exception as e:
        print(f"  [ERROR] brobrolab playwright: {e}")
    return results


def parse_dabro(base_url: str, macbook_url: str) -> list[dict]:
    soup = fetch(macbook_url)
    if not soup:
        return []
    results = []
    model_pages = _collect_macbook_model_links(soup, base_url, "remont-macbook-")
    if not model_pages:
        model_pages = _collect_macbook_model_links(soup, base_url, "macbook")
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        tables = psoup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            for row in rows:
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    repair_text = cells[0].get_text(strip=True)
                    repair = normalize_repair(repair_text)
                    if not repair:
                        continue
                    quality = normalize_quality(repair_text)
                    price_div = cells[1].find("div", class_="price-time")
                    if price_div:
                        price_line = price_div.get_text(strip=True)
                        price_line = price_line.split("Ремонт")[0].strip()
                        price_line = price_line.split("Заказать")[0].strip()
                    else:
                        price_line = cells[1].get_text(strip=True)
                    price = extract_price(price_line)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                            "quality": quality,
                        })
    return results


def parse_modmac(base_url: str, macbook_url: str) -> list[dict]:
    soup = fetch(macbook_url)
    if not soup:
        return []
    results = []
    all_links = [a for a in soup.find_all('a', href=True) if '/services/macbook/' in a['href']]
    model_pages = {}
    for a in all_links:
        href = a['href']
        last_segment = href.rstrip('/').split('/')[-1]
        if 'zamena' in last_segment or 'zamena-' in last_segment:
            continue
        link_text = a.get_text(strip=True)
        model = normalize_model_macbook(link_text)
        if model and href:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href
    target = _get_target_models()
    for model, page_url in model_pages.items():
        if model not in target:
            continue
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue
        price_blocks = psoup.find_all("div", class_="services_list_price_block")
        for pb in price_blocks:
            row = pb.parent.parent if pb.parent else None
            if not row:
                continue
            info = row.find("div", class_="services_list_info")
            if info:
                a_tag = info.find("a")
                name = a_tag.get_text(strip=True) if a_tag else info.get_text(strip=True)
                name = re.sub(r'\s+', ' ', name).strip()
            else:
                continue
            price_new = pb.find("div", class_="services_list_price_new")
            price_el = price_new if price_new else pb.find("div", class_="services_list_price")
            price_text = price_el.get_text(strip=True) if price_el else ""
            repair = normalize_repair(name)
            quality = normalize_quality(name)
            price = extract_price(price_text)
            if repair and price and price > 100:
                results.append({
                    "model": model,
                    "repair": repair,
                    "price": price,
                    "quality": quality,
                })
        results.extend(_parse_model_page_tables(page_url, model))
    return results


MOSDISPLAY_FILE_PATH = r"C:\devin_ai_powershell_devin\competitor_scraper_irepair\ЦЕНЫ Mosdisplay"


def _load_mosdisplay_workbook():
    import os
    if os.path.exists(MOSDISPLAY_FILE_PATH + ".xlsx"):
        import openpyxl
        return openpyxl.load_workbook(MOSDISPLAY_FILE_PATH + ".xlsx", data_only=True)
    if os.path.exists(MOSDISPLAY_FILE_PATH + ".xls"):
        try:
            import openpyxl
            return openpyxl.load_workbook(MOSDISPLAY_FILE_PATH + ".xls", data_only=True)
        except Exception:
            pass
        try:
            import xlrd
            return xlrd.open_workbook(MOSDISPLAY_FILE_PATH + ".xls")
        except Exception:
            pass
    return None


def _iter_ws_rows(ws, is_xlrd=False):
    if is_xlrd:
        for rx in range(ws.nrows):
            yield [ws.cell_value(rx, cx) for cx in range(ws.ncols)]
    else:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            yield [str(c.value or "").strip() for c in row]


def parse_mosdisplay(_) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("  [INFO] openpyxl not available for mosdisplay")
        return []

    results = []
    try:
        wb = _load_mosdisplay_workbook()
        if wb is None:
            print("  [WARN] mosdisplay: file not found (ЦЕНЫ Mosdisplay.xlsx/.xls)")
            return []
        is_xlrd = hasattr(wb, 'sheet_by_index')
        sheet_names = wb.sheet_names() if is_xlrd else wb.sheetnames

        for si, sheet_name in enumerate(sheet_names):
            sn_lower = sheet_name.lower().strip()
            is_macbook = "macbook" in sn_lower
            if not is_macbook:
                continue
            ws = wb.sheet_by_index(si) if is_xlrd else wb[sheet_name]

            col_repair_map = {}
            for vals in _iter_ws_rows(ws, is_xlrd):
                series_text = str(vals[0]).strip() if vals else ""
                if series_text and ("серия" in series_text.lower() or "pro" in series_text.lower() or "air" in series_text.lower()):
                    col_repair_map = {}
                    for i in range(4, min(len(vals), 20)):
                        cell_val = str(vals[i]).strip() if i < len(vals) else ""
                        if cell_val:
                            repair = normalize_repair(cell_val)
                            if repair:
                                col_repair_map[i] = repair
                    continue
                if not col_repair_map:
                    continue
                model_text = str(vals[2]).strip() if len(vals) > 2 else ""
                marking = str(vals[3]).strip() if len(vals) > 3 else ""
                model = normalize_model_macbook(marking) or normalize_model_macbook(model_text)
                if not model:
                    continue
                for col_idx, repair in col_repair_map.items():
                    cell_val = str(vals[col_idx]).strip() if col_idx < len(vals) else ""
                    if not cell_val or cell_val in ["-", "", "Уточнять", "Нужно уточнить", "Недоступно"]:
                        continue
                    first_price = cell_val.split("/")[0].split("вместе")[0].split("отдельно")[0].strip()
                    first_price = re.split(r'\s', first_price)[0].strip()
                    quality = normalize_quality(model_text) or normalize_quality(cell_val)
                    price = extract_price(first_price)
                    if price and price > 100:
                        results.append({
                            "model": model,
                            "repair": repair,
                            "price": price,
                            "quality": quality,
                        })
        if not is_xlrd:
            wb.close()
    except Exception as e:
        print(f"  [ERROR] mosdisplay xlsx: {e}")
    return results


def parse_isupport(index_url: str) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [INFO] playwright not available for isupport")
        return []

    results = []
    base = "https://www.isupport.ru"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()

            page.goto(index_url.rstrip("/") + "/", timeout=60000)
            page.wait_for_timeout(4000)

            category_links = []
            links = page.query_selector_all("a")
            for a in links:
                href = a.get_attribute("href") or ""
                text = (a.inner_text() or "").strip()
                if href and "repair-macbook/" in href and href.rstrip("/").split("/")[-1] not in ["macbook", ""]:
                    if not href.startswith("http"):
                        href = base + href
                    category_links.append({"text": text, "href": href})

            if not category_links:
                for a in links:
                    href = a.get_attribute("href") or ""
                    text = (a.inner_text() or "").strip()
                    if href and "/repair/repair-macbook/" in href:
                        if not href.startswith("http"):
                            href = base + href
                        category_links.append({"text": text, "href": href})

            model_pages = {}
            for cat in category_links:
                try:
                    page.goto(cat["href"], timeout=30000)
                except Exception:
                    continue
                page.wait_for_timeout(3000)

                cat_links = page.query_selector_all("a")
                for a in cat_links:
                    href = a.get_attribute("href") or ""
                    text = (a.inner_text() or "").strip()
                    if not href:
                        continue
                    if not href.startswith("http"):
                        href = base + href
                    if "/repair/repair-macbook/" in href and href != cat["href"]:
                        last_seg = href.rstrip("/").split("/")[-1]
                        if "macbook" in last_seg.lower() and last_seg not in ["macbook-pro", "macbook-air", "macbook"]:
                            from normalize import normalize_model_macbook_multi
                            models = normalize_model_macbook_multi(text)
                            if models:
                                for m in models:
                                    if m not in model_pages:
                                        model_pages[m] = href

            target = _get_target_models()
            for model, page_url in sorted(model_pages.items()):
                if model not in target:
                    continue
                try:
                    page.goto(page_url, timeout=30000)
                except Exception:
                    continue
                page.wait_for_timeout(3000)
                try:
                    for _ in range(8):
                        page.evaluate("window.scrollBy(0, 600)")
                        page.wait_for_timeout(200)
                except Exception:
                    pass
                page.wait_for_timeout(1000)

                content = page.content()
                soup = BeautifulSoup(content, "lxml")
                items = soup.find_all("div", class_="service-item")
                if not items:
                    items = soup.find_all("div", class_="service-price-block")
                for item in items:
                    try:
                        title_el = item.find("div", class_="service-item-title")
                        price_el = item.find("div", class_="service-item-price")
                        if not title_el or not price_el:
                            continue
                        title = title_el.get_text(strip=True)
                        price_text = price_el.get_text(strip=True)
                        repair = normalize_repair(title)
                        quality = normalize_quality(title)
                        price = extract_price(price_text)
                        if repair and price and price > 100:
                            results.append({
                                "model": model,
                                "repair": repair,
                                "price": price,
                                "quality": quality,
                            })
                    except Exception:
                        pass

            browser.close()
    except Exception as e:
        print(f"  [ERROR] isupport playwright: {e}")
    return results


JABUKA_MACBOOK_MODEL_ORDER = [
    'MacBook Air 11-13" 2010-2017',
    'MacBook Air 13" 2018-2020',
    'MacBook Air 13" M1',
    'MacBook Air 13" M2',
    'MacBook Air 13" M3',
    'MacBook Air 15" M2',
    'MacBook Air 15" M3',
    'MacBook Pro 13" 2012-2015',
    'MacBook Pro 13" 2016-2017',
    'MacBook Pro 13" 2018-2020',
    'MacBook Pro 13" M1',
    'MacBook Pro 13" M2',
    'MacBook Pro 15" 2012-2015',
    'MacBook Pro 15" 2016-2017',
    'MacBook Pro 15" 2018-2019',
    'MacBook Pro 16" Intel',
    'MacBook Pro 14" M1',
    'MacBook Pro 14" M2',
    'MacBook Pro 14" M3',
    'MacBook Pro 14" M4',
    'MacBook Pro 16" M1',
    'MacBook Pro 16" M2',
    'MacBook Pro 16" M3',
    'MacBook Pro 16" M4',
    'MacBook 12"',
]


def parse_jabuka(url: str) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [INFO] playwright not available for jabuka")
        return []

    results = []
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, timeout=25000)
            page.wait_for_timeout(5000)
            content = page.content()
            browser.close()

        soup = BeautifulSoup(content, "lxml")
        recs = soup.find_all("div", class_="t-rec", attrs={"data-record-type": "396"})

        model_list = [m for m in JABUKA_MACBOOK_MODEL_ORDER if m in _get_target_models()]

        price_blocks = []
        for rec in recs:
            text = rec.get_text(separator=" | ", strip=True)
            if "₽" in text and len(text) > 50:
                price_blocks.append(text)

        for i, block_text in enumerate(price_blocks):
            if i >= len(model_list):
                break
            model = model_list[i]
            parts = block_text.split(" | ")
            j = 0
            while j + 1 < len(parts):
                svc = parts[j].strip()
                val = parts[j + 1].strip()
                j += 2
                if not re.match(r'[Зз]амена|Ремонт', svc):
                    continue
                if '₽' not in val and val != '-':
                    continue
                if val in ('Звоните', '-'):
                    continue
                price = extract_price(val)
                if not price or price <= 100:
                    continue
                repair = normalize_repair(svc)
                quality = normalize_quality(svc)
                if repair:
                    results.append({"model": model, "repair": repair, "price": price, "quality": quality})
    except Exception as e:
        print(f"  [ERROR] jabuka: {e}")
    return results


APPLEPIE_MACBOOK_SLUG_MAP = {
    "pro16": 'MacBook Pro 16" Intel',
    "pro14": 'MacBook Pro 14" M4',
    "pro15": 'MacBook Pro 15" 2018-2019',
    "pro13": 'MacBook Pro 13" M2',
    "air15": 'MacBook Air 15" M3',
    "air13": 'MacBook Air 13" M3',
    "air11": 'MacBook Air 11-13" 2010-2017',
    "macbook12": 'MacBook 12"',
    "12": 'MacBook 12"',
}


def parse_applepie(base_url: str, macbook_url: str) -> list[dict]:
    soup = fetch(macbook_url)
    if not soup:
        return []
    results = []
    target = _get_target_models()

    price_links = [a for a in soup.find_all('a', href=True) if '/price-' in a['href']]
    model_pages = {}
    for a in price_links:
        href = a['href']
        slug = href.split('/price-')[-1].rstrip('/').split('?')[0].split('#')[0]
        model = APPLEPIE_MACBOOK_SLUG_MAP.get(slug)
        if not model:
            link_text = a.get_text(strip=True)
            model = normalize_model_macbook(link_text)
        if model and model in target:
            if not href.startswith("http"):
                href = base_url.rstrip("/") + "/" + href.lstrip("/")
            if model not in model_pages:
                model_pages[model] = href

    for model, page_url in model_pages.items():
        time.sleep(0.3)
        psoup = fetch(page_url)
        if not psoup:
            continue

        slots = psoup.find_all("div", class_="ms-slot--param-slot")
        for slot in slots:
            first_blk = slot.find("div", class_="blk_text")
            if not first_blk:
                continue
            service_name = first_blk.get_text(strip=True)

            yellow_boxes = slot.find_all("div", class_="blk_box_self")
            price_text = None
            for yb in yellow_boxes:
                style = yb.get("style", "") or yb.get("bg", "")
                if "ffd832" not in style.lower() and "#ffd832" not in style.lower():
                    bg_attr = yb.get("bg", "")
                    if "ffd832" not in bg_attr.lower():
                        continue
                bd = yb.find("div", class_="blk-data")
                if not bd:
                    continue
                txt = bd.get_text(strip=True)
                if "руб" in txt.lower() or re.match(r'^[\d\s]+руб', txt):
                    price_text = txt
                    break

            if not price_text:
                for yb in yellow_boxes:
                    bd = yb.find("div", class_="blk-data")
                    if bd:
                        txt = bd.get_text(strip=True)
                        p = extract_price(txt)
                        if p and p > 500 and not any(
                            kw in txt.lower()
                            for kw in ["мин", "час", "день", "бесплатно", "уточнять"]
                        ):
                            price_text = txt
                            break

            repair = normalize_repair(service_name)
            quality = normalize_quality(service_name)
            price = extract_price(price_text) if price_text else None
            if repair and price and price > 100:
                results.append({
                    "model": model,
                    "repair": repair,
                    "price": price,
                    "quality": quality,
                })

    return results


def parse_google_sheets(url: str) -> list[dict]:
    gid_match = re.search(r"gid=(\d+)", url)
    base = url.split("/pubhtml")[0].split("/pub?")[0].split("/pub")[0]
    if gid_match:
        csv_url = f"{base}/pub?gid={gid_match.group(1)}&single=true&output=csv"
    else:
        csv_url = f"{base}/pub?output=csv"
    try:
        resp = SESSION.get(csv_url, timeout=30)
        resp.encoding = resp.apparent_encoding or "utf-8"
        text = resp.text
        if len(text) < 20:
            print(f"  [WARN] google sheets: empty response ({len(text)} bytes)")
            return []
    except Exception as e:
        print(f"  [ERROR] google sheets: {e}")
        return []

    results = []
    reader = csv.reader(io.StringIO(text))
    header_row = None
    header_idx = -1
    for i, row in enumerate(reader):
        if not row:
            continue
        first = row[0].strip().lower() if row else ""
        if first in ["модель", "model"] or ("замена" in " ".join(row).lower() and header_row is None):
            header_row = row
            header_idx = i
            continue
        if header_row is None:
            continue
        model_text = row[0].strip() if row else ""
        if not model_text:
            continue
        model = normalize_model_macbook(model_text)
        if not model:
            continue
        for j in range(1, min(len(row), len(header_row))):
            header = header_row[j].strip() if j < len(header_row) else ""
            if not header:
                continue
            repair = normalize_repair(header)
            if not repair:
                continue
            cell = row[j].strip() if j < len(row) else ""
            prices = cell.split("/")
            for p in prices:
                price = extract_price(p.strip())
                if price and price > 100:
                    quality = normalize_quality(header) or normalize_quality(model_text)
                    results.append({
                        "model": model,
                        "repair": repair,
                        "price": price,
                        "quality": quality,
                    })
                    break
    return results
